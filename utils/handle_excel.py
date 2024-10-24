import openpyxl
import openpyxl.workbook



#定义一个读取excel的类
class ReadExcel():
    def __init__(self,file_path):
        #读取excel
        self.workbook=openpyxl.load_workbook(file_path)
        #获取excel中第一个页签
        self.worksheet=self.workbook.active
    def get_data(self):
    #eg row:(<Cell 'Students'.A1>, <Cell 'Students'.B1>, <Cell 'Students'.C1>, <Cell 'Students'.D1>, <Cell 'Students'.E1>, <Cell 'Students'.F1>, <Cell 'Students'.G1>)
    #元组里每一个元素都是一个单元格对象
        data=[]
        for row in self.worksheet.iter_rows():
            row_value=[]
            for cell in row:
                row_value.append(cell.value)
            data.append(row_value)
        return data
    
#定义一个写入Excel的类
class WriteExcel():
    #file_path:要写入文件 data：要写入数据
    def __init__(self,file_path,data):
        self.file_path=file_path
        self.data=data
        #新建一个工作本 注意：Workbook() w要大写
        self.workbook=openpyxl.Workbook()
        #获取第一页
        self.worksheet=self.workbook.active

    def write_to_excel(self):
        #一行一行的写入表中
        for row in self.data:
            self.worksheet.append(row)
        #保存excel
        self.workbook.save(self.file_path)

#以主程序来运行
if __name__=='__main__':
    import datetime
    #实例化对象
    #read_excel_obj=ReadExcel('D:/BaiduNetdiskDownload/vue3项目\StudentManagement/StudentManagement/1年1班学生信息.xlsx')
    #print(read_excel_obj.get_data())

    data=['1年1班', '张桂香', 'G586476201508266228', 'F', datetime.datetime(2015, 8, 26, 0, 0), '15239175647', '河南省刚县合川严路J座 180114'],
    write_excel_obj=WriteExcel('D:/BaiduNetdiskDownload/vue3项目\StudentManagement/StudentManagement/1年1班学生信息测试.xlsx',data)
    write_excel_obj.write_to_excel()

